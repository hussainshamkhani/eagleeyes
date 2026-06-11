import sys
import os
import io
import asyncio
from datetime import datetime

# Ensure the project directory is on the path
sys.path.insert(0, r"C:\Users\hussain\Desktop\projects\eagle-eyes")

from db.mongo import mongo_client
from api.routes import update_alert_edits, export_alerts, AlertEditsUpdateRequest

async def run_tests():
    await mongo_client.connect()
    db = mongo_client.db
    
    print("Fetching an alert from MongoDB...")
    alert = await db.alerts.find_one({})
    if not alert:
        print("No alerts found. Loading demo dataset first...")
        # Since we want to load demo data, we can call it directly or seed
        # Let's seed by calling the demo function
        from api.routes import load_demo_data
        await load_demo_data(db=db)
        # Evaluate a transaction to generate an alert
        from api.routes import evaluate_transaction_pipeline
        txn = await db.transactions.find_one({})
        if not txn:
            print("No transactions found, cannot proceed.")
            await mongo_client.disconnect()
            return
        from data.models import Transaction
        await evaluate_transaction_pipeline(Transaction(**txn), db=db)
        alert = await db.alerts.find_one({})
        assert alert is not None, "Failed to generate alert after seeding"

    alert_id = alert["alert_id"]
    ref_no = alert["ref_no"]
    print(f"Found Alert ID: {alert_id} for Transaction Ref: {ref_no}")

    # 1. Test update_alert_edits function
    print("Testing update_alert_edits function...")
    test_comment = "Compliance officer manual check passed. Verified relationship."
    test_status = "CLEAR"
    
    payload = AlertEditsUpdateRequest(comment=test_comment, user_status=test_status)
    updated_alert = await update_alert_edits(alert_id=alert_id, payload=payload, db=db)
    
    assert updated_alert.comment == test_comment, f"Expected comment {test_comment}, got {updated_alert.comment}"
    assert updated_alert.user_status == test_status, f"Expected user_status {test_status}, got {updated_alert.user_status}"
    print("update_alert_edits verified successfully!")

    # 2. Test export_alerts function
    print("Testing export_alerts function...")
    response = await export_alerts(db=db)
    
    assert response.status_code == 200, f"Expected status code 200, got {response.status_code}"
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"Got media_type {response.media_type}"
    
    # Read the streaming response body
    excel_bytes = b""
    async for chunk in response.body_iterator:
        excel_bytes += chunk
        
    print(f"Excel file size: {len(excel_bytes)} bytes")
    assert len(excel_bytes) > 0, "Excel file is empty"

    # 3. Parse workbook with openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(excel_bytes))
    sheet = wb["Flagged Transactions"]
    
    expected_headers = [
        "ref_no", "date", "branch", "sender name", "sender ID", "nationality", 
        "recipient name", "recipient country", "amount", "amount_kd", 
        "risk_score", "risk_level", "agent recommended_action", "user status", 
        "last-modified timestamp", "arize_trace_id", "comment"
    ]
    
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers: {headers}")
    assert headers == expected_headers, f"Expected: {expected_headers}, Got: {headers}"
    
    # Check that updated alert values are in the sheet
    found_alert = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == ref_no:
            found_alert = True
            print(f"Row match: {row}")
            assert row[13] == test_status, f"Expected user status {test_status}, got {row[13]}"
            assert row[16] == test_comment, f"Expected comment {test_comment}, got {row[16]}"
            
    assert found_alert, f"Alert with ref_no {ref_no} was not found in the exported sheet"
    print("All backend validations passed successfully!")
    
    await mongo_client.disconnect()

if __name__ == "__main__":
    asyncio.run(run_tests())
